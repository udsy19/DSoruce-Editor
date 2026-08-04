#!/usr/bin/env python3
"""Regenerate docs/reference/qbiq/spec/workbook-spec.json from the real qbiq .xlsx.

Resolves the workbook's CUSTOM indexedColors palette to #RRGGBB, maps every
embedded image to its real xl/media/* filename via the drawing rels, and
captures every formula as literal text (data_only=False).
"""
import json, os, re, sys, zipfile, hashlib
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.utils import get_column_letter
from PIL import Image

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))
XLSX = os.path.join(ROOT, "docs/reference/qbiq/extracted/Quantity Takeoff Sample/"
                          "Quantity Takeoff Sample/Quantity Takeoff - Formal - modern.xlsx")
OUT = os.path.join(ROOT, "docs/reference/qbiq/spec")
MEDIA_OUT = os.path.join(OUT, "media")
os.makedirs(MEDIA_OUT, exist_ok=True)

NS = {
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
}
EMU_PER_PX = 9525

z = zipfile.ZipFile(XLSX)
names = z.namelist()

# ---------------------------------------------------------------- media -----
media = {}
for n in names:
    if n.startswith("xl/media/"):
        data = z.read(n)
        base = os.path.basename(n)
        with open(os.path.join(MEDIA_OUT, base), "wb") as f:
            f.write(data)
        e = {"bytes": len(data), "sha256_16": hashlib.sha256(data).hexdigest()[:16]}
        try:
            with Image.open(os.path.join(MEDIA_OUT, base)) as im:
                e["px"] = list(im.size)
                e["format"] = im.format
                e["mode"] = im.mode
        except Exception as ex:
            e["error"] = str(ex)
        media[base] = e

# --------------------------------------------------- custom color palette ---
styles = z.read("xl/styles.xml").decode()
m = re.search(r"<indexedColors>(.*?)</indexedColors>", styles, re.S)
INDEXED = ["#" + c[2:].upper() for c in re.findall(r'rgb="([0-9a-fA-F]{8})"', m.group(1))] if m else []


def resolve(tok):
    """'indexed12' -> '#FFDC60'; passthrough for everything else."""
    if not isinstance(tok, str):
        return tok
    mm = re.fullmatch(r"indexed(\d+)", tok)
    if mm:
        i = int(mm.group(1))
        return INDEXED[i] if i < len(INDEXED) else tok
    if re.fullmatch(r"[0-9A-Fa-f]{8}", tok):
        return "#" + tok[2:].upper()
    return tok


def color_token(c):
    if c is None:
        return None
    t = getattr(c, "type", None)
    if t == "rgb" and isinstance(c.rgb, str):
        return resolve(c.rgb)
    if t == "indexed":
        return resolve(f"indexed{c.indexed}")
    if t == "theme":
        return f"theme{c.theme}+tint{round(c.tint or 0, 3)}"
    return None


# ----------------------------------------- sheet -> drawing -> media map ----
def sheet_drawing_media(sheet_part):
    """[(anchor_dict, media_basename)] for one worksheet part path."""
    rels_path = f"xl/worksheets/_rels/{os.path.basename(sheet_part)}.rels"
    if rels_path not in names:
        return []
    rels = ET.fromstring(z.read(rels_path))
    drawing = None
    for r in rels.findall("rel:Relationship", NS):
        if r.get("Type", "").endswith("/drawing"):
            drawing = os.path.normpath(os.path.join("xl/worksheets", r.get("Target"))).replace("\\", "/")
    if not drawing or drawing not in names:
        return []
    drels_path = f"{os.path.dirname(drawing)}/_rels/{os.path.basename(drawing)}.rels"
    rid2media = {}
    if drels_path in names:
        for r in ET.fromstring(z.read(drels_path)).findall("rel:Relationship", NS):
            rid2media[r.get("Id")] = os.path.basename(r.get("Target"))
    root = ET.fromstring(z.read(drawing))
    out = []
    for anch in list(root):
        tag = anch.tag.split("}")[1]
        blip = anch.find(".//a:blip", NS)
        if blip is None:
            continue
        rid = blip.get(f"{{{NS['r']}}}embed")
        d = {"anchorKind": tag}
        frm = anch.find("xdr:from", NS)
        if frm is not None:
            col = int(frm.find("xdr:col", NS).text); row = int(frm.find("xdr:row", NS).text)
            d["from"] = {"col": col, "row": row, "cell": f"{get_column_letter(col+1)}{row+1}",
                         "colOff_emu": int(frm.find("xdr:colOff", NS).text),
                         "rowOff_emu": int(frm.find("xdr:rowOff", NS).text)}
        to = anch.find("xdr:to", NS)
        if to is not None:
            col = int(to.find("xdr:col", NS).text); row = int(to.find("xdr:row", NS).text)
            d["to"] = {"col": col, "row": row, "cell": f"{get_column_letter(col+1)}{row+1}",
                       "colOff_emu": int(to.find("xdr:colOff", NS).text),
                       "rowOff_emu": int(to.find("xdr:rowOff", NS).text)}
        ext = anch.find("xdr:ext", NS)
        if ext is not None:
            cx, cy = int(ext.get("cx")), int(ext.get("cy"))
            d["ext_emu"] = {"cx": cx, "cy": cy}
            d["ext_px"] = {"w": round(cx / EMU_PER_PX), "h": round(cy / EMU_PER_PX)}
        out.append((d, rid2media.get(rid)))
    return out


# ------------------------------------------------------------- workbook -----
wb = openpyxl.load_workbook(XLSX, data_only=False)
wbxml = ET.fromstring(z.read("xl/workbook.xml"))
wbrels = {r.get("Id"): r.get("Target")
          for r in ET.fromstring(z.read("xl/_rels/workbook.xml.rels")).findall("rel:Relationship", NS)}
sheet_part = {}
for sh in wbxml.findall("main:sheets/main:sheet", NS):
    tgt = wbrels[sh.get(f"{{{NS['r']}}}id")]
    sheet_part[sh.get("name")] = "xl/" + tgt.lstrip("/")

spec = {
    "$schema_note": "Regenerated from the real qbiq workbook by Agent A. "
                    "All colors resolved from the workbook's CUSTOM xl/styles.xml <indexedColors> palette.",
    "source_file": os.path.basename(XLSX),
    "source_bytes": os.path.getsize(XLSX),
    "openpyxl": openpyxl.__version__,
    "sheet_count": len(wb.sheetnames),
    "sheet_order": wb.sheetnames,
    "indexed_palette": {f"indexed{i}": c for i, c in enumerate(INDEXED)},
    "media": media,
    "sheets": {},
}

for name in wb.sheetnames:
    ws = wb[name]
    s = {
        "index": wb.sheetnames.index(name),
        "part": sheet_part.get(name),
        "dimensions": ws.dimensions,
        "max_row": ws.max_row,
        "max_col": ws.max_column,
        "gridlines": bool(ws.sheet_view.showGridLines),
        "showRowColHeaders": bool(ws.sheet_view.showRowColHeaders),
        "state": ws.sheet_state,
        "freeze_panes": ws.freeze_panes,
        "merged": sorted(str(x) for x in ws.merged_cells.ranges),
        "col_widths": {}, "col_hidden": [], "row_heights": {},
        "images": [], "data_validations": [], "cells": {},
    }
    for k, d in ws.column_dimensions.items():
        if d.width is not None:
            s["col_widths"][k] = round(d.width, 2)
        if d.hidden:
            s["col_hidden"].append(k)
    for k, d in ws.row_dimensions.items():
        if d.height is not None:
            s["row_heights"][str(k)] = round(d.height, 2)

    for anch, mb in sheet_drawing_media(s["part"] or ""):
        e = dict(anch)
        e["media"] = mb
        if mb in media:
            e["media_px"] = media[mb].get("px")
        s["images"].append(e)

    for dv in ws.data_validations.dataValidation:
        s["data_validations"].append({
            "type": dv.type, "formula1": dv.formula1, "formula2": dv.formula2,
            "allow_blank": dv.allow_blank, "showDropDown": dv.showDropDown,
            "ranges": [str(r) for r in dv.sqref.ranges],
        })

    for row in ws.iter_rows():
        for c in row:
            e = {}
            v = c.value
            if isinstance(v, str) and v.startswith("="):
                e["f"] = v
            elif v is not None:
                e["v"] = v
            if c.number_format and c.number_format != "General":
                e["numfmt"] = c.number_format
            fd = {}
            f = c.font
            if f:
                if f.bold: fd["b"] = True
                if f.italic: fd["i"] = True
                if f.size and float(f.size) != 11: fd["sz"] = float(f.size)
                if f.name and f.name != "Calibri": fd["name"] = f.name
                col = color_token(f.color)
                if col and col not in ("#000000",): fd["color"] = col
            if fd: e["font"] = fd
            fl = c.fill
            if fl is not None and fl.patternType:
                fg = color_token(fl.fgColor)
                if fg: e["fill"] = fg; e["fillPattern"] = fl.patternType
            ad = {}
            al = c.alignment
            if al:
                if al.horizontal: ad["h"] = al.horizontal
                if al.vertical and al.vertical != "bottom": ad["v"] = al.vertical
                if al.wrap_text: ad["wrap"] = True
            if ad: e["align"] = ad
            bd = {}
            b = c.border
            if b:
                for side in ("left", "right", "top", "bottom"):
                    sv = getattr(b, side)
                    if sv is not None and sv.style:
                        bd[side] = sv.style
            if bd: e["border"] = bd
            # Skip cells that carry nothing but the sheet-wide white wash.
            if not e: continue
            if set(e) <= {"fill", "fillPattern", "border", "font"} and e.get("fill") == "#FFFFFF":
                continue
            s["cells"][c.coordinate] = e
    spec["sheets"][name] = s

with open(os.path.join(OUT, "workbook-spec.json"), "w") as f:
    json.dump(spec, f, indent=1, default=str)

print(f"sheets={len(wb.sheetnames)} palette={len(INDEXED)} media={len(media)}")
for n in wb.sheetnames:
    s = spec["sheets"][n]
    imgs = ", ".join(f"{i['media']}@{i.get('from',{}).get('cell','?')}" for i in s["images"][:3])
    print(f"  {n:30s} {s['dimensions']:10s} grid={str(s['gridlines']):5s} "
          f"imgs={len(s['images']):2d} dv={len(s['data_validations'])} cells={len(s['cells']):4d}  {imgs}")
